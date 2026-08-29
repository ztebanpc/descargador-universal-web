import sys
import os
import shutil
import time
import requests
import re
import json
import zipfile
import threading
import subprocess
import random
from concurrent.futures import ThreadPoolExecutor, as_completed

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

from flask import Flask, render_template, request, jsonify, send_from_directory
import yt_dlp

try:
    import google.generativeai as genai
    GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
    if GEMINI_API_KEY:
        genai.configure(api_key=GEMINI_API_KEY)
except Exception:
    genai = None

try:
    from PIL import Image
except ImportError:
    Image = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOAD_DIR = os.path.join(BASE_DIR, 'downloads')
AUDIO_DIR = os.path.join(DOWNLOAD_DIR, 'audio')
PINTEREST_DIR = os.path.join(DOWNLOAD_DIR, 'pinterest')
EXCEL_DIR = os.path.join(DOWNLOAD_DIR, 'excel')
ORDENES_DIR = os.path.join(DOWNLOAD_DIR, 'ordenes')

for d in [AUDIO_DIR, PINTEREST_DIR, EXCEL_DIR, ORDENES_DIR]:
    os.makedirs(d, exist_ok=True)

def sanitize_filename(name):
    clean = re.sub(r'[^\w\-_.]', '_', name)
    return clean[:40].strip('_') or f"proyecto_{int(time.time())}"

# ==========================================
# 🎵 AUDIO DOWNLOADER BLINDADO (BYPASS 403 & SABR)
# ==========================================
def process_audio_download(query, quality="192", format_type="mp3", speed="1.0", count=1, target_folder=AUDIO_DIR):
    try:
        cnt = int(count) if (str(count).isdigit() and int(count) > 0) else 1
        print(f"[Audio] Buscando {cnt} audios para: {query} | Calidad: {quality}kbps | Formato: {format_type} | Velocidad: {speed}x")
        
        postprocessors = [{
            'key': 'FFmpegExtractAudio',
            'preferredcodec': 'mp3' if format_type == 'mp3' else format_type,
            'preferredquality': quality if quality in ('128', '192', '256', '320') else '192',
        }]
        
        postprocessor_args = []
        if speed == "1.06":
            postprocessor_args = ['-filter:a', 'asetrate=46746,aresample=44100']
        elif speed and speed != "1.0":
            try:
                s_val = float(speed)
                postprocessor_args = ['-filter:a', f'atempo={s_val}']
            except ValueError:
                pass
                
        batch_id = int(time.time())
        filename_base = f"audio_{batch_id}_{random.randint(1000,9999)}"
        out_template = os.path.join(target_folder, f"{filename_base}_%(id)s.%(ext)s") if cnt > 1 else os.path.join(target_folder, f"{filename_base}.%(ext)s")
        
        # Formatos 140 / m4a / 18 evitan al 100% los bloqueos 403 de YouTube
        ydl_opts = {
            'format': '140/ba[ext=m4a]/18/bestaudio/best',
            'postprocessors': postprocessors,
            'extractor_args': {
                'youtube': {
                    'player_client': ['ios', 'android', 'web'],
                }
            },
            'outtmpl': out_template,
            'noplaylist': True,
            'quiet': True,
            'no_warnings': True,
            'max_downloads': cnt,
            'socket_timeout': 15,
        }
        if postprocessor_args:
            ydl_opts['postprocessor_args'] = {'ffmpeg': postprocessor_args}
            
        search_query = query if query.startswith('http') else f"ytsearch{cnt}:{query}"
            
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            try:
                ydl.download([search_query])
            except yt_dlp.utils.MaxDownloadsReached:
                pass
            except Exception as e:
                print(f"[Audio] Reintentando con cliente alternativo: {e}")
                # Fallback alternativo si el primer intento dio error
                ydl_opts['format'] = 'bestaudio/best'
                ydl_opts['extractor_args'] = {'youtube': {'player_client': ['mweb', 'tv', 'android']}}
                try:
                    with yt_dlp.YoutubeDL(ydl_opts) as ydl2:
                        ydl2.download([search_query])
                except Exception:
                    pass
                
        saved_files = []
        for f in os.listdir(target_folder):
            if f.startswith(filename_base) and f.endswith(('.mp3', '.m4a', '.wav', '.flac')):
                saved_files.append(f)
                
        if not saved_files:
            files = [f for f in os.listdir(target_folder) if f.endswith(('.mp3', '.m4a', '.wav', '.flac'))]
            if files:
                latest = max([os.path.join(target_folder, f) for f in files], key=os.path.getctime)
                saved_files.append(os.path.basename(latest))

        print(f"[Audio] [OK] Completados {len(saved_files)} audios.")
        return saved_files
    except Exception as e:
        print(f"[Audio] [ERROR] Error en descarga de audio: {e}")
        return []

# ==========================================
# 🎬 VIDEO SHORTS / CLIPS DOWNLOADER (Vertical HD)
# ==========================================
def process_video_download(term, count=1, format_type='9:16', target_folder=ORDENES_DIR):
    try:
        cnt = int(count) if int(count) > 0 else 1
        print(f"[Video] Buscando {cnt} clips para: {term} (Formato: {format_type})")
        
        hashtag = "#shorts " if format_type in ('9:16', 'vertical', 'any') else ""
        query = term if term.startswith('http') else f"ytsearch{cnt}:{hashtag}{term}"
        
        filename_base = f"video_{int(time.time())}_{random.randint(1000,9999)}"
        out_template = os.path.join(target_folder, f"{filename_base}_%(id)s.%(ext)s")
        
        ydl_opts = {
            'format': '18/bestvideo[ext=mp4]+bestaudio[ext=m4a]/best[ext=mp4]/best',
            'extractor_args': {
                'youtube': {
                    'player_client': ['ios', 'android', 'web'],
                }
            },
            'outtmpl': out_template,
            'noplaylist': True,
            'quiet': True,
            'no_warnings': True,
            'max_downloads': cnt,
            'socket_timeout': 15,
        }
        
        saved_files = []
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            try:
                ydl.download([query])
            except yt_dlp.utils.MaxDownloadsReached:
                pass
            except Exception as e:
                print(f"[Video] Reintentando video con cliente mweb: {e}")
                ydl_opts['extractor_args'] = {'youtube': {'player_client': ['mweb', 'android']}}
                try:
                    with yt_dlp.YoutubeDL(ydl_opts) as ydl2:
                        ydl2.download([query])
                except Exception:
                    pass
                
        for f in os.listdir(target_folder):
            if f.startswith(filename_base) and f.endswith(('.mp4', '.mov', '.webm', '.mkv')):
                saved_files.append(f)
                
        print(f"[Video] [OK] Descargados {len(saved_files)} clips de video.")
        return saved_files
    except Exception as e:
        print(f"[Video] [ERROR]: {e}")
        return []

def process_bulk_audio(links, speed="1.0"):
    print(f"[Excel] Procesando {len(links)} enlaces en lote...")
    for link in links:
        process_audio_download(link, speed=speed, count=1, target_folder=EXCEL_DIR)
        time.sleep(0.8)
    print(f"[Excel] [OK] Todos los enlaces de Excel fueron procesados.")

# ==========================================
# 📌 PINTEREST SCRAPER LIMPIO (SIN ICONOS BASURA)
# ==========================================
def extract_pinterest_raw_links(query, count=20):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }
    
    videos = []
    images = []
    
    clean_q = re.sub(r'(\b\d+\b|\bfotos?\b|\bde\b|\bpara\b|\bvertical\b|\bhorizontal\b|\bcuadrada\b)', '', query, flags=re.IGNORECASE).strip()
    if not clean_q: clean_q = query.strip()
    
    if clean_q.startswith('http') or 'pin.it' in clean_q or 'pinterest.com' in clean_q:
        url = clean_q if clean_q.startswith('http') else 'https://' + clean_q
        try:
            res = requests.get(url, headers=headers, allow_redirects=True, timeout=10)
            if res.status_code == 200:
                html = res.text
                v_matches = re.findall(r'https://[^"\'\s<>]+?\.pinimg\.com/[^"\'\s<>]+\.mp4', html)
                for v in v_matches:
                    if v not in videos: videos.append(v)
                i_matches = re.findall(r'https://i\.pinimg\.com/[^"\'\s<>]+?\.(?:jpg|jpeg|png|webp)', html)
                for m in i_matches:
                    if any(bad in m for bad in ['avatar', '75x75', '200x150', '150x150', '30x30', '256x256', 'icon']):
                        continue
                    orig = re.sub(r'/(?:236x|474x|564x|736x)/', '/originals/', m)
                    if orig not in images:
                        images.append(orig)
        except Exception as e:
            print("[Pinterest] Error al resolver enlace:", e)
    else:
        try:
            range_req = max(30, count * 2)
            cmd = [sys.executable, "-m", "gallery_dl", "--get-urls", "--range", f"1-{range_req}", f"https://www.pinterest.com/search/pins/?q={clean_q.replace(' ', '+')}"]
            out = subprocess.check_output(cmd, text=True, errors="ignore", timeout=20)
            for line in out.splitlines():
                u = line.strip().replace('| ', '').strip()
                if u.startswith('ytdl:'):
                    u = u.replace('ytdl:', '')
                if u.lower().split('?')[0].endswith(('.mp4', '.mov', '.webm')):
                    if u not in videos: videos.append(u)
                elif u.startswith('http'):
                    if any(bad in u for bad in ['avatar', '75x75', '200x150', '150x150', '30x30', '256x256', 'icon']):
                        continue
                    orig = re.sub(r'/(?:236x|474x|564x|736x)/', '/originals/', u)
                    if orig not in images:
                        images.append(orig)
        except Exception as e:
            print("[Pinterest] Error en búsqueda temática con gallery-dl:", e)

    return {"videos": videos, "images": images}

def download_single_image(u, batch_folder, idx, format_type):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }
    try:
        r = requests.get(u, headers=headers, timeout=8)
        if r.status_code == 200:
            ext = u.split('?')[0].split('.')[-1].lower()
            if ext not in ('jpg', 'jpeg', 'png', 'webp'): ext = 'jpg'
            temp_name = f"temp_{idx}_{random.randint(100,999)}.{ext}"
            temp_fp = os.path.join(batch_folder, temp_name)
            with open(temp_fp, 'wb') as f:
                f.write(r.content)
                
            if Image:
                try:
                    with Image.open(temp_fp) as img_pil:
                        w, h = img_pil.size
                    # Descartar iconos pequeños / placeholders basura
                    if w < 300 or h < 300:
                        os.remove(temp_fp); return None
                    fmt = format_type.lower()
                    if fmt in ('9:16', 'vertical') and h < w * 1.15:
                        os.remove(temp_fp); return None
                    elif fmt in ('16:9', 'horizontal') and w < h * 1.15:
                        os.remove(temp_fp); return None
                    elif fmt in ('1:1', 'cuadrado') and not (0.85 <= w/h <= 1.15):
                        os.remove(temp_fp); return None
                except Exception:
                    pass
                    
            final_fname = f"pin_foto_{idx+1}.{ext}"
            final_fp = os.path.join(batch_folder, final_fname)
            shutil.move(temp_fp, final_fp)
            return final_fname
    except Exception:
        pass
    return None

def process_pinterest_download(query, format_type='any', media_type='both', pin_tab='individual', count=1, target_folder=PINTEREST_DIR, create_zip=True):
    try:
        req_count = int(count) if (str(count).isdigit() and int(count) > 0) else 1
        print(f"[Pinterest] Procesando: {query} | Modo: {pin_tab} | Formato: {format_type} | Tipo: {media_type} | Cantidad: {req_count}")
        raw = extract_pinterest_raw_links(query, count=req_count)
        videos = raw.get("videos", [])
        images = raw.get("images", [])
        
        # Si el usuario pidió solo videos en Pinterest y Pinterest no tiene MP4s para esa temática, buscar clips HD directamente
        if media_type == 'videos' and len(videos) == 0 and not query.startswith('http'):
            print("[Pinterest] Fallback automático a extractor de clips de video para:", query)
            vids = process_video_download(query, count=req_count, format_type=format_type, target_folder=target_folder)
            zip_filename = None
            if create_zip and len(vids) > 1:
                zip_filename = f"pinterest_videos_{int(time.time())}.zip"
                zip_path = os.path.join(target_folder, zip_filename)
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for fname in vids:
                        fpath = os.path.join(target_folder, fname)
                        if os.path.exists(fpath):
                            zipf.write(fpath, fname)
            return vids, zip_filename
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        
        saved_files = []
        batch_id = int(time.time())
        batch_folder = os.path.join(target_folder, f"lote_{batch_id}")
        os.makedirs(batch_folder, exist_ok=True)
        
        if pin_tab == 'individual':
            if videos:
                v_url = videos[0]
                try:
                    r = requests.get(v_url, headers=headers, stream=True, timeout=12)
                    if r.status_code == 200:
                        fname = f"pin_video_{batch_id}.mp4"
                        fp = os.path.join(batch_folder, fname)
                        with open(fp, 'wb') as f:
                            for chunk in r.iter_content(chunk_size=8192):
                                f.write(chunk)
                        saved_files.append(fname)
                except Exception as e:
                    print("[Pinterest] Error guardando video individual:", e)
            elif images:
                i_url = images[0]
                try:
                    r = requests.get(i_url, headers=headers, timeout=10)
                    if r.status_code == 200:
                        ext = i_url.split('?')[0].split('.')[-1].lower()
                        if ext not in ('jpg', 'jpeg', 'png', 'webp'): ext = 'jpg'
                        fname = f"pin_foto_{batch_id}.{ext}"
                        fp = os.path.join(batch_folder, fname)
                        with open(fp, 'wb') as f:
                            f.write(r.content)
                        saved_files.append(fname)
                except Exception as e:
                    print("[Pinterest] Error guardando foto individual:", e)
                    
            if saved_files:
                for f in saved_files:
                    shutil.copy(os.path.join(batch_folder, f), os.path.join(target_folder, f))
                return saved_files, None

        limit = req_count
        if pin_tab == 'tablero' and req_count == 1 and len(images) > 1:
            limit = len(images)
            
        items_to_download = []
        if media_type in ('both', 'videos'):
            for v in videos:
                items_to_download.append({'type': 'video', 'url': v})
        if media_type in ('both', 'photos'):
            for img in images:
                items_to_download.append({'type': 'image', 'url': img})
                
        image_urls = [it['url'] for it in items_to_download if it['type'] == 'image'][:limit]
        video_urls = [it['url'] for it in items_to_download if it['type'] == 'video'][:limit]
        
        with ThreadPoolExecutor(max_workers=6) as executor:
            future_to_idx = {executor.submit(download_single_image, u, batch_folder, i, format_type): i for i, u in enumerate(image_urls)}
            for future in as_completed(future_to_idx):
                res_fname = future.result()
                if res_fname:
                    saved_files.append(res_fname)
                    
        for idx, v_url in enumerate(video_urls):
            if len(saved_files) >= limit: break
            try:
                r = requests.get(v_url, headers=headers, stream=True, timeout=12)
                if r.status_code == 200:
                    fname = f"pin_video_{batch_id}_{idx+1}.mp4"
                    fp = os.path.join(batch_folder, fname)
                    with open(fp, 'wb') as f:
                        for chunk in r.iter_content(chunk_size=8192):
                            f.write(chunk)
                    saved_files.append(fname)
            except Exception:
                pass

        for f in saved_files:
            shutil.copy(os.path.join(batch_folder, f), os.path.join(target_folder, f))
            
        zip_filename = None
        if create_zip and len(saved_files) > 1:
            zip_filename = f"pinterest_descarga_{batch_id}.zip"
            zip_path = os.path.join(target_folder, zip_filename)
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for fname in saved_files:
                    fpath = os.path.join(batch_folder, fname)
                    if os.path.exists(fpath):
                        zipf.write(fpath, fname)
                    
        return saved_files, zip_filename
    except Exception as e:
        print("[Pinterest] [ERROR]:", e)
        return [], None

# ==========================================
# 🧠 CEREBRO DE INTERPRETACIÓN INTELIGENTE CON ALTA DIVERSIDAD CREATIVA
# ==========================================
SONG_BANKS = {
    "tristeza": [
        "Maná - Te Lloré Un Río", "Adele - Someone Like You", "Lewis Capaldi - Someone You Loved",
        "Sam Smith - Stay With Me", "Billie Eilish - Lovely", "Gary Jules - Mad World",
        "Coldplay - Fix You", "Evanescence - My Immortal", "Olivia Rodrigo - Drivers License",
        "Ed Sheeran - Happier", "Sia - Chandelier Piano Version", "Kodaline - All I Want",
        "Duncan Laurence - Arcade", "Lord Huron - The Night We Met", "Tom Odell - Another Love"
    ],
    "triste": [
        "Maná - Te Lloré Un Río", "Adele - Someone Like You", "Lewis Capaldi - Someone You Loved",
        "Sam Smith - Stay With Me", "Billie Eilish - Lovely", "Gary Jules - Mad World",
        "Tom Odell - Another Love", "Coldplay - The Scientist", "Passenger - Let Her Go"
    ],
    "cumpleaños": [
        "Cepillín - Las Mañanitas", "Parchis - Cumpleaños Feliz", "Stevie Wonder - Happy Birthday",
        "Canción Infantil - Feliz en tu Día", "The Beatles - Birthday", "Katy Perry - Birthday",
        "Vicente Fernández - Las Mañanitas", "Raffaella Carrà - Hay Que Venir Al Sur"
    ],
    "fiesta": [
        "Bad Bunny - Tití Me Preguntó", "Feid - Feliz Cumpleaños Ferxxo", "Don Omar - Danza Kuduro",
        "Daddy Yankee - Gasolina", "Rauw Alejandro - Todo De Ti", "J Balvin - Mi Gente",
        "Bizarrap & Quevedo - Bzrp Music Sessions", "Farruko - Pepas", "Pitbull - Fireball"
    ],
    "aesthetic": [
        "Harry Styles - As It Was", "The Weeknd - Blinding Lights", "Glass Animals - Heat Waves",
        "Dua Lipa - Levitating", "Steve Lacy - Bad Habit", "The Neighbourhood - Sweater Weather",
        "Arctic Monkeys - I Wanna Be Yours", "Beach House - Space Song", "Mac DeMarco - Chamber of Reflection"
    ],
    "desamor": [
        "Shakira - Monotonía", "Morat - No Se Va", "Rauw Alejandro - 2/Catorce",
        "Camilo - Ropa Cara", "Sebastian Yatra - Devuélveme El Corazón", "Reik - Ya Me Enteré"
    ]
}

def clean_term_string(s):
    s = re.sub(r'(\b\d+\b|\b(tres|dos|cuatro|cinco|diez|veinte|treinta|vertical(?:es)?|horizontal(?:es)?|cuadrad[ao]s?|formato|velocidad|calidad|mp3|wav|flac|de|fotos?|im[aá]genes|videos?|vídeos?|audios?|cancion(?:es)?|m[uú]sicas?|anticopyright|anti copyright|quiero|necesito|me gustaría|me gustaria|unos|unas|un|una|para|que tengan que ver con|que tengan|relacionados? con|tem[aá]ticas?)\b)', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[:;.,_()\[\]\\0-9]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def normalize_llm_items(items):
    normalized = []
    for itm in items:
        itype = itm.get('type', 'photo').lower()
        if 'audio' in itype or 'music' in itype or 'sound' in itype:
            itype = 'audio'
        elif 'video' in itype or 'clip' in itype:
            itype = 'video'
        else:
            itype = 'photo'

        term = itm.get('term', '')
        if not term:
            if itm.get('artist') and itm.get('title'):
                term = f"{itm['artist']} - {itm['title']}"
            elif itm.get('title'):
                term = itm['title']
            elif itm.get('description'):
                term = itm['description']
            elif itm.get('query'):
                term = itm['query']
            elif itm.get('name'):
                term = itm['name']
                
        term = str(term).strip()
        if not term: continue

        fmt = str(itm.get('format', 'any')).lower()
        if '9:16' in fmt or 'vertical' in fmt:
            fmt = '9:16'
        elif '16:9' in fmt or 'horizontal' in fmt:
            fmt = '16:9'
        elif '1:1' in fmt or 'cuadrad' in fmt:
            fmt = '1:1'
        elif itype == 'audio':
            fmt = 'mp3'
        else:
            fmt = 'any'

        speed = str(itm.get('speed', '1.0')).replace('x', '')
        if speed not in ('1.0', '1.06', '1.25', '1.5', '2.0'):
            speed = '1.06' if ('1.06' in str(itm.get('speed', '')) or 'anticopyright' in str(itm).lower()) else '1.0'

        count = int(itm.get('count', 1))
        if itype == 'audio': count = 1
        elif count <= 0: count = 5

        normalized.append({
            "term": term,
            "type": itype,
            "format": fmt,
            "speed": speed,
            "count": count
        })
    return normalized

def interpret_ai_order_with_gemini_37(prompt_text):
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if api_key and genai:
        try:
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel(
                'gemini-3.7-flash',
                generation_config={"temperature": 0.88, "top_p": 0.95}
            )
            system_instruction = """
Eres el Director Creativo de Edición Audiovisual y Productor Musical para creadores de contenido (TikTok, Reels, Shorts).
Tu misión es interpretar la solicitud del usuario con inteligencia profunda, VARIEDAD CREATIVA CONSTANTE y desglosarla en recursos de máxima calidad.

REGLAS DE ORO:
1. DIVERSIDAD: VARÍA SIEMPRE las recomendaciones y canciones. No repitas siempre las mismas canciones estándar. Elige temas famosos, virales, himnos dramáticos o hits modernos según el matiz exacto del pedido.
2. AUDIOS: Si el usuario pide un número de canciones de cierta temática (ej: "3 músicas de tristeza", "2 canciones de fiesta", "canción romántica"):
   - DEBES generar EXACTAMENTE esa cantidad de líneas individuales de audio.
   - Cada línea DEBE ser una canción real, famosa con su "Artista - Título".
   - NUNCA pongas nombres genéricos como "tres tristezas". Siempre canciones reales.
   - Cantidad siempre es 1 por cada canción individual.
   - Si el usuario mencionó 1.06x o anticopyright, pon speed: "1.06", de lo contrario "1.0".
3. VIDEOS: Extrae los términos de búsqueda visual más potentes para B-Rolls y clips. Asigna cantidad y formato ("9:16", "16:9").
4. FOTOS: Extrae el término estético limpio, cantidad y formato ("1:1", "9:16", "16:9", "any").
5. RECOMENDACIÓN: Explica brevemente en 1 frase por qué seleccionaste esas canciones y cómo potencian el ritmo y la emoción de la edición.

Devuelve ÚNICAMENTE un JSON válido con esta estructura:
{
  "recommendation": "Explicación breve de por qué se eligieron esas canciones específicas y el estilo visual",
  "project_name": "nombre_corto_del_proyecto",
  "items": [
    {
      "term": "Artista - Título exacto O término de búsqueda visual",
      "type": "photo" | "video" | "audio",
      "format": "1:1" | "9:16" | "16:9" | "any" | "mp3",
      "speed": "1.0" | "1.06" | "1.25",
      "count": 1
    }
  ]
}
"""
            res = model.generate_content(system_instruction + "\n\nSolicitud del usuario:\n" + prompt_text)
            clean_res = res.text.strip().replace('```json', '').replace('```', '').strip()
            parsed = json.loads(clean_res)
            raw_items = parsed.get("items", [])
            if raw_items:
                norm_items = normalize_llm_items(raw_items)
                if norm_items:
                    print(f"[Gemini 3.7 Flash] Orden interpretada con éxito ({len(norm_items)} items).")
                    return {
                        "recommendation": parsed.get("recommendation", "Orden estructurada con inteligencia creativa."),
                        "project_name": sanitize_filename(parsed.get("project_name", "proyecto_edicion")),
                        "items": norm_items
                    }
        except Exception as e:
            print("[Gemini 3.7 Flash] Fallback a motor semántico local:", e)
            
    return parse_order_with_deep_context(prompt_text)

def parse_order_with_deep_context(prompt_text):
    prompt_clean = prompt_text.strip()
    slug = re.sub(r'[^\w\-_]', '_', prompt_clean[:25]).strip('_') or "proyecto_edicion"
    
    normalized = re.sub(r'(\d+)\.(\d+)(?:x)?', r'\1_DOT_\2', prompt_clean)
    normalized = re.sub(r'(?<=[^\s,;])\s+(\d+\s+(?:fotos?|im[aá]genes|videos?|vídeos?|audios?|cancion(?:es)?|m[uú]sicas?))', r', \1', normalized, flags=re.IGNORECASE)
    normalized = re.sub(r'(?<=[^\s,;])\s+((?:tres|dos|cuatro|cinco|diez|veinte)\s+(?:fotos?|im[aá]genes|videos?|vídeos?|audios?|cancion(?:es)?|m[uú]sicas?))', r', \1', normalized, flags=re.IGNORECASE)
    
    segments = re.split(r'[,;\n]|\by\b|\be\b|\bcon\b', normalized, flags=re.IGNORECASE)
    global_speed = "1.06" if any(w in prompt_clean.lower() for w in ["1.06", "1.6", "anti copyright", "anticopyright"]) else "1.0"
    
    items = []
    recommended_songs = []
    word_to_num = {'un': 1, 'una': 1, 'uno': 1, 'dos': 2, 'tres': 3, 'cuatro': 4, 'cinco': 5, 'seis': 6, 'diez': 10, 'veinte': 20, 'treinta': 30}
    
    for seg in segments:
        seg_clean_dots = seg.replace('_DOT_', '.').strip()
        if not seg_clean_dots or len(seg_clean_dots) < 2: continue
        
        seg_lower = seg_clean_dots.lower()
        seg_without_speed = re.sub(r'\b1\.06(?:x)?\b|\b1\.25(?:x)?\b|\b1\.5(?:x)?\b|\b2\.0(?:x)?\b|\b1\.0(?:x)?\b', '', seg_lower).strip()
        
        count = 0
        count_match = re.search(r'\b(\d+)\b', seg_without_speed)
        if count_match:
            count = int(count_match.group(1))
        else:
            for w, n in word_to_num.items():
                if re.search(rf'\b{w}\b', seg_without_speed):
                    count = n; break
                    
        is_audio = any(w in seg_lower for w in ['audio', 'cancion', 'canción', 'musica', 'música', 'mp3', 'sound', 'tema', 'track', 'tristeza', 'triste', 'desamor', 'mana', 'harry style'])
        is_video = any(w in seg_lower for w in ['video', 'vídeo', 'clip', 'reels', 'shorts', 'broll', 'b-roll']) and not is_audio
        is_photo = any(w in seg_lower for w in ['foto', 'imagen', 'imágenes', 'portada', 'wallpaper', 'pic', 'meme', 'memes']) and not is_audio and not is_video
        
        if any(w in seg_lower for w in ['harry style', 'harry styles', 'bad bunny', 'duki', 'feid', 'taylor swift', 'morat', 'queen', 'beatles', 'mana', 'maná', 'adele']):
            is_audio = True; is_video = False; is_photo = False
            
        fmt = '9:16' if ('9:16' in seg_lower or 'vertical' in seg_lower) else ('16:9' if ('16:9' in seg_lower or 'horizontal' in seg_lower) else ('1:1' if ('1:1' in seg_lower or 'cuadrad' in seg_lower or 'meme' in seg_lower) else 'any'))
        speed = "1.06" if any(w in seg_lower for w in ["1.06", "1.6", "anticopyright", "anti copyright"]) else global_speed
        
        clean = clean_term_string(seg_without_speed)
        if not clean or len(clean) < 2: clean = seg_clean_dots.strip()
        
        if is_audio:
            audio_count = count if count > 0 else 1
            matched_theme = None
            for theme_key in SONG_BANKS:
                if theme_key in seg_lower:
                    matched_theme = theme_key; break
                    
            if matched_theme:
                bank = SONG_BANKS[matched_theme]
                sample_k = min(audio_count, len(bank))
                selected_samples = random.sample(bank, sample_k)
                for song_name in selected_samples:
                    recommended_songs.append(song_name)
                    items.append({
                        "term": song_name,
                        "type": "audio",
                        "format": "mp3",
                        "speed": speed,
                        "count": 1
                    })
            else:
                items.append({
                    "term": clean.title(),
                    "type": "audio",
                    "format": "mp3",
                    "speed": speed,
                    "count": 1
                })
        elif is_video:
            v_count = count if count > 0 else 5
            if 'cumpleaños' in clean.lower():
                clean = "Cumpleaños fiesta celebración confeti"
            items.append({
                "term": clean,
                "type": "video",
                "format": fmt if fmt != 'any' else '9:16',
                "speed": "1.0",
                "count": v_count
            })
        else: # Foto
            p_count = count if count > 0 else 10
            if 'meme' in clean.lower():
                clean = "Memes graciosos virales"
                fmt = "1:1"
            elif 'cumpleaños' in clean.lower():
                clean = "Cumpleaños decoración aesthetic"
            items.append({
                "term": clean,
                "type": "photo",
                "format": fmt,
                "speed": "1.0",
                "count": p_count
            })
            
    recommendation = f"Se estructuraron {len(items)} recursos creativos. "
    if recommended_songs:
        recommendation += f"Se seleccionaron las canciones ({', '.join(recommended_songs[:3])}) para darle la vibra musical exacta a la edición."
    else:
        recommendation += "Recursos listos con formato y velocidad balanceados para tu proyecto."
        
    return {
        "recommendation": recommendation,
        "project_name": slug,
        "items": items
    }

# ==========================================
# 🌐 RUTAS FLASK
# ==========================================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/health')
def health():
    return jsonify({"status": "ok", "service": "descargador-universal-web"}), 200

@app.route('/api/download_audio', methods=['POST'])
def download_audio():
    data = request.get_json(silent=True) or request.form
    query = data.get('query', '').strip()
    quality = data.get('quality', '192')
    format_type = data.get('format', 'mp3')
    speed = data.get('speed', '1.0')
    count = data.get('cantidad', 1)
    
    if not query:
        return jsonify({"status": "error", "message": "Debes ingresar un enlace o nombre.", "note": "Verifica que el término de búsqueda no esté vacío."}), 400
        
    saved_files = process_audio_download(query, quality=quality, format_type=format_type, speed=speed, count=count)
    if saved_files:
        file_urls = [f"/api/files/audio/{f}" for f in saved_files]
        zip_url = None
        if len(saved_files) > 1:
            batch_id = int(time.time())
            zip_filename = f"audios_lote_{batch_id}.zip"
            zip_path = os.path.join(AUDIO_DIR, zip_filename)
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for f in saved_files:
                    fp = os.path.join(AUDIO_DIR, f)
                    if os.path.isfile(fp):
                        zipf.write(fp, f)
            zip_url = f"/api/files/audio/{zip_filename}"
            
        return jsonify({
            "status": "success",
            "message": f"Descargados {len(saved_files)} audios exitosamente ({speed}x).",
            "total": len(saved_files),
            "files": file_urls,
            "download_url": file_urls[0] if len(file_urls) == 1 else None,
            "zip_url": zip_url
        })
    return jsonify({"status": "error", "message": "No se pudo extraer el audio.", "note": "Comprueba que el video o búsqueda esté disponible públicamente."}), 500

@app.route('/api/download_excel', methods=['POST'])
def download_excel():
    file = request.files.get('file')
    speed = request.form.get('speed', '1.0')
    if not file:
        return jsonify({"status": "error", "message": "No se subió ningún archivo.", "note": "Selecciona un archivo .xlsx o .csv válido."}), 400

    links = []
    fname = file.filename.lower()
    try:
        if fname.endswith('.csv'):
            import csv, io
            content = file.read().decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(content))
            for row in reader:
                for cell in row:
                    c = cell.strip()
                    if c.startswith('http'):
                        links.append(c)
        elif fname.endswith('.xlsx') or fname.endswith('.xls'):
            if openpyxl is None:
                return jsonify({"status": "error", "message": "Librería openpyxl no instalada."}), 500
            import io
            wb = openpyxl.load_workbook(io.BytesIO(file.read()))
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell and str(cell).strip().startswith('http'):
                            links.append(str(cell).strip())
        else:
            return jsonify({"status": "error", "message": "Formato no soportado. Usa .xlsx o .csv"}), 400

        if not links:
            return jsonify({"status": "error", "message": "No se encontraron enlaces en el archivo.", "note": "Asegúrate de que las celdas contengan URLs completas que empiecen por http://"}), 400

        threading.Thread(target=process_bulk_audio, args=(links, speed)).start()
        return jsonify({"status": "success", "message": f"Procesando {len(links)} audios desde Excel a {speed}x en segundo plano."})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error leyendo archivo: {e}"}), 500

@app.route('/api/download_pinterest', methods=['POST'])
def download_pinterest():
    data = request.get_json(silent=True) or request.form
    query = data.get('query', '').strip()
    format_type = data.get('format', 'any')
    media_type = data.get('media_type', 'both')
    pin_tab = data.get('pin_tab', 'individual')
    count = data.get('cantidad', 1)
    
    if not query:
        return jsonify({"status": "error", "message": "Debes ingresar una URL o término.", "note": "Pega un enlace de pin/tablero o escribe una palabra clave para buscar."}), 400
        
    saved, zip_name = process_pinterest_download(query, format_type=format_type, media_type=media_type, pin_tab=pin_tab, count=count)
    if saved:
        file_urls = [f"/api/files/pinterest/{f}" for f in saved]
        zip_url = f"/api/files/pinterest/{zip_name}" if zip_name else None
        return jsonify({
            "status": "success",
            "message": f"Descargados {len(saved)} archivos en calidad original.",
            "total": len(saved),
            "files": file_urls,
            "download_url": file_urls[0] if len(file_urls) == 1 else None,
            "zip_url": zip_url
        })
    return jsonify({
        "status": "error", 
        "message": "No se encontraron archivos con los criterios indicados.",
        "note": "Si seleccionaste un formato específico (ej. 9:16 o 16:9), es posible que los pines encontrados no coincidan con esa proporción. Prueba cambiando a 'Cualquiera' o ajusta el término de búsqueda."
    }), 404

@app.route('/api/ai_order_assistant', methods=['POST'])
def ai_order_assistant():
    data = request.get_json(silent=True) or request.form
    prompt = data.get('prompt', '').strip()
    
    if not prompt:
        return jsonify({"status": "error", "message": "Por favor ingresa o dicta una orden."}), 400
        
    res = interpret_ai_order_with_gemini_37(prompt)
    return jsonify({
        "status": "success",
        "recommendation": res.get("recommendation", "Orden analizada y optimizada con Gemini 3.7 Flash."),
        "project_name": res.get("project_name", "proyecto_edicion"),
        "items": res.get("items", [])
    })

def execute_single_item(itm, project_folder):
    term = itm.get('term', '').strip()
    itype = itm.get('type', 'photo')
    fmt = itm.get('format', 'any')
    speed = itm.get('speed', '1.0')
    cnt = int(itm.get('count', 1))
    
    if not term: return []
    
    try:
        if itype == 'audio':
            return process_audio_download(term, speed=speed, count=cnt, target_folder=project_folder)
        elif itype == 'video':
            return process_video_download(term, count=cnt, format_type=fmt, target_folder=project_folder)
        else: # photo
            pins, _ = process_pinterest_download(term, format_type=fmt, media_type='photos', pin_tab='tematica', count=cnt, target_folder=project_folder, create_zip=False)
            return pins or []
    except Exception as e:
        print(f"[Item Execution Error] {term}: {e}")
        return []

@app.route('/api/execute_custom_order', methods=['POST'])
def execute_custom_order():
    data = request.get_json(silent=True) or request.form
    raw_name = data.get('project_name', 'Ordenes_IA')
    project_slug = sanitize_filename(raw_name)
    items = data.get('items', [])
    
    if not items:
        return jsonify({"status": "error", "message": "No hay elementos en la orden."}), 400
        
    project_folder_name = f"Ordenes_IA_{project_slug}_{int(time.time())}"
    project_folder = os.path.join(ORDENES_DIR, project_folder_name)
    os.makedirs(project_folder, exist_ok=True)
    downloaded_files = []
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        future_to_item = {executor.submit(execute_single_item, itm, project_folder): itm for itm in items}
        for future in as_completed(future_to_item):
            try:
                res_files = future.result()
                if res_files:
                    downloaded_files.extend(res_files)
            except Exception as e:
                print("[Task Exception]:", e)
                
    master_zip_name = f"Ordenes_IA_{project_slug}.zip"
    master_zip_path = os.path.join(ORDENES_DIR, master_zip_name)
    
    with zipfile.ZipFile(master_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(project_folder):
            for file in files:
                if not file.endswith('.zip'):
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, project_folder)
                    zipf.write(file_path, arcname)
                    
    file_urls = [f"/api/files/ordenes/{f}" for f in downloaded_files]
    return jsonify({
        "status": "success",
        "message": f"Proyecto completado con {len(downloaded_files)} archivos descargados.",
        "total": len(downloaded_files),
        "download_url": file_urls[0] if len(file_urls) == 1 else None,
        "zip_url": f"/api/files/ordenes/{master_zip_name}",
        "files": file_urls
    })

@app.route('/api/files/<category>/<filename>')
def serve_file(category, filename):
    folder_map = {
        'audio': AUDIO_DIR,
        'pinterest': PINTEREST_DIR,
        'excel': EXCEL_DIR,
        'ordenes': ORDENES_DIR
    }
    target_dir = folder_map.get(category, DOWNLOAD_DIR)
    return send_from_directory(target_dir, filename, as_attachment=True)

if __name__ == '__main__':
    print("[Sistema] Iniciando Descargador Universal...")
    app.run(host='0.0.0.0', port=10000, debug=False)
